import pandas as pd
import logging
from typing import Dict, List, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import inspect
import uuid
from datetime import datetime
import io
from io import BytesIO

logger = logging.getLogger(__name__)


class DataImporter:
    def __init__(self, db: Session):
        self.db = db

    def import_csv(self, file_content: bytes, model_class, column_mapping: Dict[str, str]) -> Dict[str, Any]:
        """Import data from CSV file"""
        try:
            text = file_content.decode("utf-8", errors="replace")  # safer than BytesIO
            df = pd.read_csv(io.StringIO(text))
            return self._import_dataframe(df, model_class, column_mapping)
        except Exception as e:
            logger.error(f"CSV import error: {e}")
            raise ValueError(f"CSV import failed: {str(e)}")

    def import_excel(
        self, file_content: bytes, model_class, column_mapping: Dict[str, str], sheet_name: Optional[str] = 0
    ) -> Dict[str, Any]:
        """Import data from Excel file"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name)
            return self._import_dataframe(df, model_class, column_mapping)
        except Exception as e:
            logger.error(f"Excel import error: {e}")
            raise ValueError(f"Excel import failed: {str(e)}")

    def _import_dataframe(self, df: pd.DataFrame, model_class, column_mapping: Dict[str, str]) -> Dict[str, Any]:
        """Process DataFrame and import data"""
        # Validate mapping
        self._validate_mapping(df.columns, column_mapping)

        # Rename columns
        df = df.rename(columns=column_mapping)

        # Normalize column names
        df.columns = [c.strip() for c in df.columns]

        # Convert DataFrame to records
        records = df.to_dict("records")

        return self._import_records(records, model_class)

    def _validate_mapping(self, file_columns: List[str], column_mapping: Dict[str, str]):
        """Ensure mapping keys exist in file"""
        for file_col in column_mapping.keys():
            if file_col not in file_columns:
                raise ValueError(f"Column '{file_col}' not found in uploaded file")

    def _import_records(self, records: List[Dict[str, Any]], model_class) -> Dict[str, Any]:
        successful = 0
        failed = 0
        errors = []

        for i, record in enumerate(records, 1):
            try:
                cleaned = self._clean_record(record, model_class)
                if not cleaned:
                    continue
                obj = model_class(**cleaned)
                self.db.add(obj)
                successful += 1
            except Exception as e:
                failed += 1
                errors.append(f"Row {i}: {str(e)}")
                logger.warning(f"Row {i} failed: {e}")

        if successful > 0:
            self.db.commit()

        return {
            "total_processed": len(records),
            "successful": successful,
            "failed": failed,
            "errors": errors,
        }

    def _clean_record(self, record: Dict[str, Any], model_class) -> Dict[str, Any]:
        cleaned = {}
        mapper = inspect(model_class)

        for key, value in record.items():
            if pd.isna(value) or value == "":
                continue

            if key not in mapper.columns:
                continue

            column = mapper.columns[key]
            cleaned[key] = self._convert_value(value, column.type)

        return cleaned

    def _convert_value(self, value, column_type):
        """Safe type conversion"""
        try:
            py_type = getattr(column_type, "python_type", str)

            if py_type == uuid.UUID:
                return uuid.UUID(str(value))

            if py_type == datetime:
                return pd.to_datetime(value, errors="coerce").to_pydatetime() if value else None

            if py_type in (int, float):
                return py_type(value)

            return py_type(value)

        except Exception as e:
            logger.warning(f"Conversion error for '{value}': {e}")
            return value


# ---- Model-specific importers ----

class FineImporter(DataImporter):
    def import_fines(
        self, file_content: bytes, file_type: str, column_mapping: Dict[str, str], sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        from app.models import Fine, Location, Vehicle
        import openpyxl
        
        try:
            if file_type == "excel":
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                
                # Use provided sheet name or auto-detect
                if sheet_name is None:
                    # Look for fines-related sheets
                    for sheet in wb.sheetnames:
                        sheet_lower = sheet.lower()
                        if any(term in sheet_lower for term in ['штраф', 'fine', 'нарушен']):
                            sheet_name = sheet
                            break
                    else:
                        sheet_name = wb.sheetnames[0]
                
                sheet = wb[sheet_name]
                print(f"Processing fines from sheet: {sheet.title}")
                
                # Find header row
                header_row = None
                header_keywords = ['issued_at', 'plate_number', 'violation_code', 'amount', 'address', 'status']
                
                for row_num in range(1, min(10, sheet.max_row + 1)):
                    header_match_count = 0
                    for col in range(1, min(10, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_num, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if any(keyword in cell_lower for keyword in header_keywords):
                                header_match_count += 1
                    
                    if header_match_count >= 2:
                        header_row = row_num
                        break
                
                first_data_row = header_row + 1 if header_row else 2
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Cache for vehicles and locations to avoid duplicate lookups
                vehicle_cache = {}  # plate_number -> vehicle_id
                location_cache = {}  # address -> location_id
                
                for row_num in range(first_data_row, sheet.max_row + 1):
                    try:
                        # Skip empty rows
                        if sheet.cell(row=row_num, column=1).value is None:
                            continue
                            
                        # Extract row data
                        row_data = {}
                        for col in range(1, sheet.max_column + 1):
                            header_cell = sheet.cell(row=header_row, column=col).value if header_row else f"col_{col}"
                            value_cell = sheet.cell(row=row_num, column=col).value
                            
                            if value_cell is not None:
                                # Apply column mapping
                                mapped_key = column_mapping.get(str(header_cell).strip(), str(header_cell).strip())
                                row_data[mapped_key] = value_cell
                        
                        # Skip if no valid data
                        if not row_data:
                            continue
                            
                        # Process vehicle (get or create)
                        plate_number = str(row_data.get('plate_number', '')).strip()
                        if not plate_number:
                            continue
                            
                        if plate_number in vehicle_cache:
                            vehicle_id = vehicle_cache[plate_number]
                        else:
                            # Look for existing vehicle
                            vehicle = self.db.query(Vehicle).filter(Vehicle.plate_number == plate_number).first()
                            if vehicle:
                                vehicle_id = vehicle.id
                            else:
                                # Create new vehicle
                                vehicle = Vehicle(
                                    id=uuid.uuid4(),
                                    plate_number=plate_number,
                                    type="car"  # default type
                                )
                                self.db.add(vehicle)
                                self.db.flush()  # Get the ID without committing
                                vehicle_id = vehicle.id
                            vehicle_cache[plate_number] = vehicle_id
                        
                        # Process location (get or create)
                        address = str(row_data.get('address', '')).strip()
                        if not address:
                            continue
                            
                        if address in location_cache:
                            location_id = location_cache[address]
                        else:
                            # Look for existing location
                            location = self.db.query(Location).filter(Location.address == address).first()
                            if location:
                                location_id = location.id
                            else:
                                # Create new location
                                location = Location(
                                    id=uuid.uuid4(),
                                    address=address
                                )
                                self.db.add(location)
                                self.db.flush()
                                location_id = location.id
                            location_cache[address] = location_id
                        
                        # Process date
                        issued_at = row_data.get('issued_at')
                        if isinstance(issued_at, datetime):
                            issued_at_date = issued_at
                        elif isinstance(issued_at, str):
                            try:
                                issued_at_date = datetime.strptime(issued_at, '%Y-%m-%d')
                            except:
                                # Try other formats
                                try:
                                    issued_at_date = datetime.strptime(issued_at, '%d.%m.%Y')
                                except:
                                    issued_at_date = datetime.now()
                        else:
                            issued_at_date = datetime.now()
                        
                        # Create fine record
                        fine_data = {
                            'vehicle_id': vehicle_id,
                            'location_id': location_id,
                            'amount': float(row_data.get('amount', 0)),
                            'issued_at': issued_at_date,
                            'violation_code': str(row_data.get('violation_code', '')),
                            'status': str(row_data.get('status', 'issued')),
                            'visibility': 'private'
                        }
                        
                        fine = Fine(**fine_data)
                        self.db.add(fine)
                        success_count += 1
                        
                        if success_count % 50 == 0:
                            self.db.commit()
                            
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
                
                self.db.commit()
                
                return {
                    "total_processed": sheet.max_row - first_data_row + 1,
                    "successful": success_count,
                    "failed": error_count,
                    "errors": errors
                }
                
            else:
                # Fall back to generic CSV import (will need similar modifications)
                raise Exception("CSV import not yet implemented for fines with relationships")
                
        except Exception as e:
            self.db.rollback()
            raise Exception(f"Fines import failed: {str(e)}")


class AccidentImporter(DataImporter):
    def import_accidents(
        self, file_content: bytes, file_type: str, column_mapping: Dict[str, str], sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        from app.models import Accident, Location
        import openpyxl
        
        try:
            if file_type == "excel":
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                
                # Use provided sheet name or auto-detect
                if sheet_name is None:
                    for sheet in wb.sheetnames:
                        sheet_lower = sheet.lower()
                        if any(term in sheet_lower for term in ['дтп', 'accident', 'incident']):
                            sheet_name = sheet
                            break
                    else:
                        sheet_name = wb.sheetnames[0]
                
                sheet = wb[sheet_name]
                print(f"Processing accidents from sheet: {sheet.title}")
                
                # Find header row
                header_row = None
                header_keywords = ['occurred_at', 'accident_type', 'severity', 'casualties', 'address']
                
                for row_num in range(1, min(10, sheet.max_row + 1)):
                    header_match_count = 0
                    for col in range(1, min(10, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_num, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if any(keyword in cell_lower for keyword in header_keywords):
                                header_match_count += 1
                    
                    if header_match_count >= 2:
                        header_row = row_num
                        break
                
                first_data_row = header_row + 1 if header_row else 2
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Cache for locations
                location_cache = {}
                
                for row_num in range(first_data_row, sheet.max_row + 1):
                    try:
                        # Skip empty rows
                        if sheet.cell(row=row_num, column=1).value is None:
                            continue
                            
                        # Extract row data
                        row_data = {}
                        for col in range(1, sheet.max_column + 1):
                            header_cell = sheet.cell(row=header_row, column=col).value if header_row else f"col_{col}"
                            value_cell = sheet.cell(row=row_num, column=col).value
                            
                            if value_cell is not None:
                                mapped_key = column_mapping.get(str(header_cell).strip(), str(header_cell).strip())
                                row_data[mapped_key] = value_cell
                        
                        if not row_data:
                            continue
                            
                        # Process location
                        address = str(row_data.get('address', '')).strip()
                        if not address:
                            continue
                            
                        if address in location_cache:
                            location_id = location_cache[address]
                        else:
                            location = self.db.query(Location).filter(Location.address == address).first()
                            if location:
                                location_id = location.id
                            else:
                                location = Location(
                                    id=uuid.uuid4(),
                                    address=address
                                )
                                self.db.add(location)
                                self.db.flush()
                                location_id = location.id
                            location_cache[address] = location_id
                        
                        # Process date
                        occurred_at = row_data.get('occurred_at')
                        if isinstance(occurred_at, datetime):
                            occurred_at_date = occurred_at
                        elif isinstance(occurred_at, str):
                            try:
                                occurred_at_date = datetime.strptime(occurred_at, '%Y-%m-%d %H:%M:%S')
                            except:
                                try:
                                    occurred_at_date = datetime.strptime(occurred_at, '%Y-%m-%d')
                                except:
                                    occurred_at_date = datetime.now()
                        else:
                            occurred_at_date = datetime.now()
                        
                        # Create accident record
                        accident_data = {
                            'location_id': location_id,
                            'accident_type': str(row_data.get('accident_type', '')),
                            'severity': str(row_data.get('severity', 'minor')),
                            'occurred_at': occurred_at_date,
                            'casualties': int(row_data.get('casualties', 0)),
                            'visibility': 'private'
                        }
                        
                        accident = Accident(**accident_data)
                        self.db.add(accident)
                        success_count += 1
                        
                        if success_count % 50 == 0:
                            self.db.commit()
                            
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
                
                self.db.commit()
                
                return {
                    "total_processed": sheet.max_row - first_data_row + 1,
                    "successful": success_count,
                    "failed": error_count,
                    "errors": errors
                }
                
            else:
                raise Exception("CSV import not yet implemented for accidents")
                
        except Exception as e:
            self.db.rollback()
            raise Exception(f"Accidents import failed: {str(e)}")


class TrafficLightImporter(DataImporter):
    def import_traffic_lights(
        self, file_content: bytes, file_type: str, column_mapping: Dict[str, str], sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        from app.models import TrafficLight, Location
        import openpyxl
        
        try:
            if file_type == "excel":
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                
                # Use provided sheet name or auto-detect
                if sheet_name is None:
                    for sheet in wb.sheetnames:
                        sheet_lower = sheet.lower()
                        if any(term in sheet_lower for term in ['светофор', 'traffic', 'light']):
                            sheet_name = sheet
                            break
                    else:
                        sheet_name = wb.sheetnames[0]
                
                sheet = wb[sheet_name]
                print(f"Processing traffic lights from sheet: {sheet.title}")
                
                # Find header row
                header_row = None
                header_keywords = ['type', 'status', 'install_date', 'address']
                
                for row_num in range(1, min(10, sheet.max_row + 1)):
                    header_match_count = 0
                    for col in range(1, min(10, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row_num, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if any(keyword in cell_lower for keyword in header_keywords):
                                header_match_count += 1
                    
                    if header_match_count >= 2:
                        header_row = row_num
                        break
                
                first_data_row = header_row + 1 if header_row else 2
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Cache for locations
                location_cache = {}
                
                for row_num in range(first_data_row, sheet.max_row + 1):
                    try:
                        # Skip empty rows
                        if sheet.cell(row=row_num, column=1).value is None:
                            continue
                            
                        # Extract row data
                        row_data = {}
                        for col in range(1, sheet.max_column + 1):
                            header_cell = sheet.cell(row=header_row, column=col).value if header_row else f"col_{col}"
                            value_cell = sheet.cell(row=row_num, column=col).value
                            
                            if value_cell is not None:
                                mapped_key = column_mapping.get(str(header_cell).strip(), str(header_cell).strip())
                                row_data[mapped_key] = value_cell
                        
                        if not row_data:
                            continue
                            
                        # Process location
                        address = str(row_data.get('address', '')).strip()
                        if not address:
                            continue
                            
                        if address in location_cache:
                            location_id = location_cache[address]
                        else:
                            location = self.db.query(Location).filter(Location.address == address).first()
                            if location:
                                location_id = location.id
                            else:
                                location = Location(
                                    id=uuid.uuid4(),
                                    address=address
                                )
                                self.db.add(location)
                                self.db.flush()
                                location_id = location.id
                            location_cache[address] = location_id
                        
                        # Process dates
                        install_date = row_data.get('install_date')
                        if isinstance(install_date, datetime):
                            install_date_val = install_date.date()
                        elif isinstance(install_date, str):
                            try:
                                install_date_val = datetime.strptime(install_date, '%Y-%m-%d').date()
                            except:
                                install_date_val = datetime.now().date()
                        else:
                            install_date_val = datetime.now().date()
                        
                        # Create traffic light record
                        traffic_light_data = {
                            'location_id': location_id,
                            'type': str(row_data.get('type', '')),
                            'status': str(row_data.get('status', 'working')),
                            'install_date': install_date_val,
                            'last_maintenance': None  # Can be updated later
                        }
                        
                        traffic_light = TrafficLight(**traffic_light_data)
                        self.db.add(traffic_light)
                        success_count += 1
                        
                        if success_count % 50 == 0:
                            self.db.commit()
                            
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
                
                self.db.commit()
                
                return {
                    "total_processed": sheet.max_row - first_data_row + 1,
                    "successful": success_count,
                    "failed": error_count,
                    "errors": errors
                }
                
            else:
                raise Exception("CSV import not yet implemented for traffic lights")
                
        except Exception as e:
            self.db.rollback()
            raise Exception(f"Traffic lights import failed: {str(e)}")


class EvacuationImporter(DataImporter):
    def import_evacuations(
        self, file_content: bytes, file_type: str, column_mapping: Dict[str, str], sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        from app.models import Evacuation, Location
        import pandas as pd
        from io import BytesIO
        
        try:
            # First, ensure we have a location
            print("Checking available locations...")
            
            existing_location = self.db.query(Location).first()
            
            if existing_location:
                print(f"Using existing location with ID: {existing_location.id}")
                location_id = existing_location.id
            else:
                print("No locations found, creating a new one...")
                
                location_data = {'id': uuid.uuid4()}
                
                try:
                    if hasattr(Location, 'name'):
                        location_data['name'] = "Default Evacuation Location"
                    elif hasattr(Location, 'address'):
                        location_data['address'] = "City-wide evacuation data"
                    elif hasattr(Location, 'description'):
                        location_data['description'] = "Default evacuation location"
                except:
                    pass
                
                try:
                    new_location = Location(**location_data)
                    self.db.add(new_location)
                    self.db.commit()
                    location_id = new_location.id
                    print(f"Created new location with ID: {location_id}")
                except Exception as e:
                    print(f"Error creating location: {e}")
                    location_id = uuid.uuid4()
                    print(f"Using generated location ID: {location_id}")
            
            if file_type == "excel":
                # Use pandas to read the Excel file with proper sheet handling
                xls = pd.ExcelFile(BytesIO(file_content))
                
                # Auto-detect the right sheet if not provided
                if sheet_name is None:
                    sheet_name = self._detect_evacuation_sheet(xls.sheet_names)
                
                print(f"Available sheets: {xls.sheet_names}")
                print(f"Selected sheet: {sheet_name}")
                
                # Read the specific sheet
                df = pd.read_excel(BytesIO(file_content), sheet_name=sheet_name)
                
                print(f"DataFrame shape: {df.shape}")
                print("Original columns:", df.columns.tolist())
                print("First few rows:")
                print(df.head())
                
                # Apply column mappings
                reverse_mapping = {v: k for k, v in column_mapping.items()}
                df_renamed = df.rename(columns=reverse_mapping)
                
                print("After applying mappings:")
                print("Mapped columns:", df_renamed.columns.tolist())
                print("Available data columns:", [col for col in df_renamed.columns if col in df_renamed])
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df_renamed.iterrows():
                    try:
                        # Skip empty rows - check if evacuated_at is NaN or None
                        evacuated_at_value = row.get('evacuated_at')
                        if pd.isna(evacuated_at_value) or evacuated_at_value is None:
                            continue
                            
                        print(f"Processing row {index}: evacuated_at = {evacuated_at_value}")
                        
                        # Parse date
                        evacuated_at = self._parse_date(evacuated_at_value)
                        if not evacuated_at:
                            print(f"Skipping row {index} - invalid date: {evacuated_at_value}")
                            continue
                        
                        # Create evacuation record with safe type conversion
                        evacuation_data = {
                            'evacuated_at': evacuated_at,
                            'towing_vehicles_count': self._safe_int(row.get('towing_vehicles_count')),
                            'dispatches_count': self._safe_int(row.get('dispatches_count')),
                            'evacuations_count': self._safe_int(row.get('evacuations_count')),
                            'revenue': self._safe_float(row.get('revenue')),
                            'location_id': location_id,
                            'visibility': 'private'
                        }
                        
                        print(f"Evacuation data: {evacuation_data}")
                        
                        evacuation = Evacuation(**evacuation_data)
                        self.db.add(evacuation)
                        success_count += 1
                        
                        if success_count % 50 == 0:
                            self.db.commit()
                            print(f"Committed {success_count} records...")
                            
                    except Exception as e:
                        error_count += 1
                        error_msg = f"Row {index}: {str(e)}"
                        errors.append(error_msg)
                        print(f"Error on row {index}: {e}")
                        import traceback
                        print(traceback.format_exc())
                        continue
                
                self.db.commit()
                print(f"Import completed: {success_count} successful, {error_count} failed")
                
                return {
                    "total_processed": len(df_renamed),
                    "successful": success_count,
                    "failed": error_count,
                    "errors": errors
                }
                
        except Exception as e:
            self.db.rollback()
            import traceback
            print(f"Import error: {str(e)}")
            print(traceback.format_exc())
            raise Exception(f"Excel import failed: {str(e)}")
    
    def _detect_evacuation_sheet(self, sheet_names):
        """Auto-detect the evacuation data sheet"""
        evacuation_keywords = ['эвакуация', 'evacuation', 'эвакуации', 'evacuations']
        
        for sheet in sheet_names:
            sheet_lower = sheet.lower()
            if any(keyword in sheet_lower for keyword in evacuation_keywords):
                return sheet
        
        # If no evacuation sheet found, look for data sheets (not route/map sheets)
        non_route_keywords = ['маршрут', 'route', 'map', 'аналитик', 'пример']
        for sheet in sheet_names:
            sheet_lower = sheet.lower()
            if not any(keyword in sheet_lower for keyword in non_route_keywords):
                return sheet
        
        # Fallback to first sheet
        return sheet_names[0] if sheet_names else None
    
    def _parse_date(self, date_value):
        """Parse date from various formats"""
        if pd.isna(date_value) or date_value is None:
            return None
            
        if isinstance(date_value, datetime):
            return date_value
        elif isinstance(date_value, pd.Timestamp):
            return date_value.to_pydatetime()
        elif isinstance(date_value, str):
            date_str = str(date_value).strip()
            # Try different date formats
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d',
                '%d.%m.%Y',
                '%m/%d/%Y',
                '%d-%m-%Y'
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
        return None
    
    def _safe_int(self, value, default=0):
        """Safely convert to integer"""
        try:
            if pd.isna(value) or value is None:
                return default
            return int(float(value))
        except (ValueError, TypeError):
            return default
    
    def _safe_float(self, value, default=0.0):
        """Safely convert to float"""
        try:
            if pd.isna(value) or value is None:
                return default
            return float(value)
        except (ValueError, TypeError):
            return default